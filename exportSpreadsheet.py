import os
import uuid
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet import cell_range
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.drawing.image import Image as OpenpyxlImage
from copy import copy
from PIL import Image as PILImage
import re
from time import sleep

# TODO:
# Make times 12 hour
# _get_variable_pos cant get merged cell coords properly
# cant iterate through multiple year groups


class ExportSpreadsheets:
    def __init__(self, fileName):
        self.scoresheetImg = None

        self.fileName = fileName
        self.workbook = Workbook()

        self.template = load_workbook(
            filename="template.xlsx", data_only=True)

        self._create_worksheet("runsheet", "runsheet")

        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

        self.templateVars = {"runsheet": self._get_variable_pos(
            self.template["runsheet"]), "scoresheet": self._get_variable_pos(self.template["scoresheet"])}

        print(self.templateVars["scoresheet"])

    def save(self):
        self.workbook.save(self.fileName)

    def get_team_data(self, fileName):
        teamData = load_workbook(filename=fileName)

        return None

    def add_data(self, data):
        for year in data:
            for game in data[year]:
                # print(game, year)
                self._add_runsheet_data(data[year][game])
                self._add_scoresheet_data(data[year][game], game, year)

    def _add_runsheet_data(self, data):
        self._add_data_to_worksheet("runsheet", data["time"] + "_c" + str(data["court"]),
                                    data["white"] + " (W) vs " + data["black"] + " (B)", "runsheet")

    def _add_scoresheet_data(self, data, key, year):
        add = self._add_data_to_worksheet

        whiteTeam = data["white"]
        blackTeam = data["black"]

        add(key, "team_white_title", whiteTeam)
        add(key, "team_black_title", blackTeam)
        add(key, "team_white", whiteTeam)
        add(key, "team_black", blackTeam)

        add(key, "court", data["court"])
        add(key, "time", data["time"]+":00")
        add(key, "year_group", "GRADES " + year)

        worksheet = self.workbook[key]

        startingYNum, startingXNum = coordinate_from_string(
            self.templateVars["scoresheet"]["a_player_num"])
        startingYName, startingXName = coordinate_from_string(
            self.templateVars["scoresheet"]["a_player_name"])

        for index, player in enumerate(data[whiteTeam]):
            worksheet[startingYNum+str(int(startingXNum)+index)] = player[1]
            worksheet[startingYName+str(int(startingXName)+index)] = player[0]

        startingYNum, startingXNum = coordinate_from_string(
            self.templateVars["scoresheet"]["b_player_num"])
        startingYName, startingXName = coordinate_from_string(
            self.templateVars["scoresheet"]["b_player_name"])

        for index, player in enumerate(data[blackTeam]):
            worksheet[startingYNum+str(int(startingXNum)+index)] = player[1]
            worksheet[startingYName+str(int(startingXName)+index)] = player[0]

    def _add_data_to_worksheet(self, sheet, var, data, type="scoresheet"):
        if sheet not in self.workbook.sheetnames:
            self._create_worksheet(type, sheet)

        worksheet = self.workbook[sheet]

        if var in self.templateVars[type]:
            worksheet[self.templateVars[type][var]] = data
        else:
            print(var)
            raise Exception("var not in template variables")

    def _get_variable_pos(self, worksheet):
        vars = {}

        for row in worksheet.iter_rows():
            for cell in row:
                cellValue = self._get_merged_cell_val(worksheet, cell)

                if not cellValue:
                    cellValue = cell.value

                if not isinstance(cellValue, str):
                    continue

                cellVar = re.search(r"_.+_", cellValue)

                if cellVar:
                    val = cellVar.group()[1:-1]
                    vars[val] = cell.coordinate

        return vars

    def _get_merged_cell_val(self, sheet, cell):
        coords = cell.coordinate

        for range_ in sheet.merged_cells.ranges:
            cellRange = str(range_).split(":")
            merged_cells = sheet[cellRange[0]:cellRange[1]][0]
            coordInRange = False

            cellStr = ""

            for cell in merged_cells:
                if cell.value:
                    cellStr += str(cell.value)

                    if cell.coordinate == coords:
                        coordInRange = True

            if "GRADES" in cellStr and coordInRange:
                return cellStr
        return None

    def _create_worksheet(self, type, name):
        worksheet = self.workbook.create_sheet(name)

        template = self.template[type]

        self._copy_sheet(template, worksheet, type)

    def _copy_sheet(self, source_sheet, target_sheet, type):
        # copy all the cel values and styles
        self._copy_cells(source_sheet, target_sheet)
        self._copy_sheet_attributes(source_sheet, target_sheet, type)

    def _copy_sheet_attributes(self, source_sheet, target_sheet, type):
        target_sheet.sheet_format = copy(source_sheet.sheet_format)
        target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
        target_sheet.merged_cells = copy(source_sheet.merged_cells)
        target_sheet.page_margins = copy(source_sheet.page_margins)
        target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

        target_sheet._images = []

        # Save the NSBL img on the template sheet
        if type == "scoresheet" and self.scoresheetImg == None:
            imgs = copy(source_sheet._images)

            for img in imgs:
                self.scoresheetImg = PILImage.open(img.ref)
                self.scoresheetImg.resize((img.width, img.height))
        elif type == "scoresheet":
            target_sheet._images = [self.scoresheetImg]

        # set row dimensions
        # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
        for rn in range(len(source_sheet.row_dimensions)):
            target_sheet.row_dimensions[rn] = copy(
                source_sheet.row_dimensions[rn])

        if source_sheet.sheet_format.defaultColWidth is None:
            print('Unable to copy default column wide')
        else:
            target_sheet.sheet_format.defaultColWidth = copy(
                source_sheet.sheet_format.defaultColWidth)

        # set specific column width and hidden property
        # we cannot copy the entire column_dimensions attribute so we copy selected attributes
        for key, value in source_sheet.column_dimensions.items():
            # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
            target_sheet.column_dimensions[key].min = copy(
                source_sheet.column_dimensions[key].min)
            # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
            target_sheet.column_dimensions[key].max = copy(
                source_sheet.column_dimensions[key].max)
            target_sheet.column_dimensions[key].width = copy(
                source_sheet.column_dimensions[key].width)  # set width for every column
            target_sheet.column_dimensions[key].hidden = copy(
                source_sheet.column_dimensions[key].hidden)

    def _copy_cells(self, source_sheet, target_sheet):
        for (row, col), source_cell in source_sheet._cells.items():
            target_cell = target_sheet.cell(column=col, row=row)

            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)


if __name__ == "__main__":
    testWorkbook = ExportSpreadsheets("test.xlsx")
    testData = {"3/4": {"9c1": {"ateams": [["name", 0], ["name", 0]], "wjdf": [
        ["name", 1], ["name", 1]], "time": "9", "white": "ateams", "black": "wjdf", "court": 1, }}, "4/5": {"11c1": {"ateams": [["name", 0], ["name", 0]], "wjdf": [
            ["name", 1], ["name", 1]], "time": "11", "white": "ateams", "black": "wjdf", "court": 1, }}, }

    testWorkbook.add_data(testData)

    print("saving data")
    testWorkbook.save()
